from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter, SearchFilter
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import TarifCotisation, Cotisation
from .serializers import TarifCotisationSerializer, CotisationSerializer


class TarifCotisationViewSet(viewsets.ModelViewSet):
    queryset = TarifCotisation.objects.all()
    serializer_class = TarifCotisationSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["annee", "categorie_affiliation"]


class CotisationViewSet(viewsets.ModelViewSet):
    queryset = Cotisation.objects.select_related("membre").all()
    serializer_class = CotisationSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_fields = ["membre", "annee", "trimestre", "statut", "mode_paiement"]
    search_fields = ["membre__nom", "membre__prenom", "membre__email"]
    ordering_fields = ["annee", "trimestre", "statut", "membre__nom", "date_paiement"]
    ordering = ["-annee", "-trimestre"]

    @action(detail=True, methods=["post"], url_path="payer")
    def payer(self, request, pk=None):
        """Enregistre un paiement pour une cotisation."""
        cotisation = self.get_object()
        montant = request.data.get("montant_paye")
        mode = request.data.get("mode_paiement", "ESPECES")
        commentaire = request.data.get("commentaire", "")

        if montant is None:
            return Response(
                {"detail": "montant_paye est requis."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            montant = float(montant)
        except (TypeError, ValueError):
            return Response(
                {"detail": "montant_paye doit être un nombre."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if montant <= 0:
            return Response(
                {"detail": "montant_paye doit être positif."},
                status=status.HTTP_400_BAD_REQUEST
            )

        cotisation.montant_paye = montant
        cotisation.mode_paiement = mode
        cotisation.date_paiement = timezone.now().date()
        if commentaire:
            cotisation.commentaire = commentaire
        cotisation.save()  # recalculer_statut() est appelé dans save()

        serializer = CotisationSerializer(cotisation)
        return Response(serializer.data)

    @action(detail=False, methods=["post"], url_path="generer-trimestre")
    def generer_trimestre(self, request):
        """Génère les cotisations pour un trimestre donné (sans Celery)."""
        from members.models import Membre

        annee = request.data.get("annee", timezone.now().year)
        trimestre = request.data.get("trimestre")

        if not trimestre:
            return Response(
                {"detail": "trimestre est requis (1-4)."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            annee = int(annee)
            trimestre = int(trimestre)
            if trimestre not in [1, 2, 3, 4]:
                raise ValueError()
        except (TypeError, ValueError):
            return Response(
                {"detail": "annee et trimestre doivent être valides."},
                status=status.HTTP_400_BAD_REQUEST
            )

        membres = Membre.objects.filter(statut_compte="ACTIF")
        crees = 0
        existants = 0

        for membre in membres:
            tarif = TarifCotisation.objects.filter(
                categorie_affiliation=membre.categorie_affiliation,
                annee=annee
            ).first()

            montant = tarif.montant_trimestriel if tarif else 100

            _, created = Cotisation.objects.get_or_create(
                membre=membre,
                annee=annee,
                trimestre=trimestre,
                defaults={
                    "montant_attendu": montant,
                    "montant_paye": 0,
                    "statut": "EN_RETARD",
                }
            )
            if created:
                crees += 1
            else:
                existants += 1

        return Response({
            "detail": f"Cotisations T{trimestre}/{annee} générées.",
            "crees": crees,
            "existants": existants,
            "total": crees + existants,
        })

    @action(detail=False, methods=["get"], url_path="resume-membre/(?P<membre_id>[0-9]+)")
    def resume_membre(self, request, membre_id=None):
        """Résumé des cotisations d'un membre."""
        cotisations = Cotisation.objects.filter(
            membre_id=membre_id
        ).order_by("-annee", "-trimestre")

        serializer = CotisationSerializer(cotisations, many=True)
        total_du = sum(c.montant_attendu for c in cotisations)
        total_paye = sum(c.montant_paye for c in cotisations)

        return Response({
            "cotisations": serializer.data,
            "total_du": float(total_du),
            "total_paye": float(total_paye),
            "solde": float(total_du - total_paye),
            "taux": round(float(total_paye) / float(total_du) * 100, 1) if total_du > 0 else 0,
        })

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cotisations"

        header_fill = PatternFill("solid", fgColor="CE1126")
        header_font = Font(bold=True, color="FFFFFF")
        headers = [
            "ID", "Membre", "Catégorie", "Année", "Trimestre",
            "Montant attendu", "Montant payé", "Solde", "Statut",
            "Mode paiement", "Date paiement"
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row, c in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=c.id)
            ws.cell(row=row, column=2, value=str(c.membre))
            ws.cell(row=row, column=3, value=c.membre.get_categorie_affiliation_display())
            ws.cell(row=row, column=4, value=c.annee)
            ws.cell(row=row, column=5, value=f"T{c.trimestre}")
            ws.cell(row=row, column=6, value=float(c.montant_attendu))
            ws.cell(row=row, column=7, value=float(c.montant_paye))
            ws.cell(row=row, column=8, value=float(c.montant_attendu - c.montant_paye))
            ws.cell(row=row, column=9, value=c.get_statut_display())
            ws.cell(row=row, column=10, value=c.get_mode_paiement_display() if c.mode_paiement else "")
            ws.cell(row=row, column=11, value=str(c.date_paiement or ""))

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="cotisations_cndd_fdd.xlsx"'
        wb.save(response)
        return response
