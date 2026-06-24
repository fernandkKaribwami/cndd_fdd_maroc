from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from django.http import HttpResponse
from django.db import transaction
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from datetime import datetime, date
import re
import unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Membre, ProfilEtudiant
from core.models import Cycle, Domaine, Filiere, Niveau
from .serializers import (
    MembreListSerializer, MembreDetailSerializer, MembreCreateUpdateSerializer,
    ProfilEtudiantSerializer
)
from .filters import MembreFilter


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────

VILLES_MAROC = [
    "Agadir", "Al Hoceima", "Azilal", "Azrou", "Béni Mellal", "Berkane",
    "Berrechid", "Casablanca", "Chefchaouen", "Dakhla", "El Jadida",
    "Errachidia", "Essaouira", "Fès", "Guelmim", "Ifrane", "Inezgane",
    "Kénitra", "Khémisset", "Khouribga", "Laâyoune", "Larache",
    "Marrakech", "Meknès", "Midelt", "Mohammedia", "Nador",
    "Ouarzazate", "Oujda", "Rabat", "Safi", "Salé", "Settat",
    "Sidi Kacem", "Sidi Slimane", "Tan-Tan", "Tanger", "Taroudant",
    "Taourirt", "Taza", "Temara", "Tétouan", "Tiznit", "Zagora",
]

# Mots-clés → code CELLULE_CHOICES
CELLULE_MAP = {
    "TANGER": "TANGER_TETOUAN_OUJDA",
    "TETOUAN": "TANGER_TETOUAN_OUJDA",
    "TETOUAN": "TANGER_TETOUAN_OUJDA",
    "OUJDA": "TANGER_TETOUAN_OUJDA",
    "KENITRA": "KENITRA",
    "AGADIR": "AGADIR",
    "RABAT": "RABAT_SALE",
    "SALE": "RABAT_SALE",
    "LAAYOUNE": "LAAYOUNE_DAKHLA",
    "DAKHLA": "LAAYOUNE_DAKHLA",
    "FES": "FEZ_MEKNES",
    "FEZ": "FEZ_MEKNES",
    "MEKNES": "FEZ_MEKNES",
    "CASABLANCA": "CASABLANCA",
    "SETTAT": "CASABLANCA",
}

CELLULE_VALID = {
    "TANGER_TETOUAN_OUJDA", "KENITRA", "AGADIR", "RABAT_SALE",
    "LAAYOUNE_DAKHLA", "FEZ_MEKNES", "CASABLANCA", "AUTRE",
}

FEMALE_FIRST_NAMES = {
    "monia", "joëlla", "joella", "sandra", "aloysie", "florence", "élodie",
    "elodie", "diane", "divine", "francine", "collade", "ornella", "lady",
    "belyse", "olivella", "patience", "angélique", "angelique", "sandrine",
    "gérardine", "gerardine", "alice", "bella", "fleur", "princia", "carelle",
    "clairia", "leila", "eunice", "joyce", "delis", "christa", "mélance",
    "melance", "mireille", "praxède", "praxede", "kerry", "kerren", "livia",
    "deltine", "diesse", "verlaine", "dominique", "fabiola", "grace",
    "marlène", "marlene", "nadège", "nadege", "noëlle", "noelle",
    "quinthia", "rose", "sylvie", "yvette", "fatima", "nimpaye",
    "nahimana", "iteka", "uwihoreye", "mado", "nkurunziza",
    "niyobuhungiro", "niyonkuru", "nsanganiye", "mukamarakiza", "niyokindi",
    "namugisha", "ishimwe", "ngabirano", "singirankabo", "ndizeye",
    "irangabiye", "matsiko", "irankunda",
}


# ─────────────────────────────────────────────────────────────────────────────
# UTILITAIRES PARTAGÉS (PDF + Excel)
# ─────────────────────────────────────────────────────────────────────────────

def _plain(text: str) -> str:
    """Supprime les accents et met en majuscules pour la comparaison floue."""
    return ''.join(
        c for c in unicodedata.normalize('NFD', text.upper())
        if unicodedata.category(c) != 'Mn'
    )


# Règles de catégorie : (regex sur texte normalisé, label_canonique, statut, cycle_nom)
_CAT_DEFS = [
    (re.compile(r'MED[EI]C[IE]NE.*(4|4E|QUATRE)', re.I), 'Médecine 4ème année et +', 'ETUDIANT', 'Médecine'),
    (re.compile(r'(4E|4EME|4\+).*MED[EI]C[IE]NE', re.I), 'Médecine 4ème année et +', 'ETUDIANT', 'Médecine'),
    (re.compile(r'MED[EI]C[IE]NE', re.I),              'Médecine 1ère–3ème année',  'ETUDIANT', 'Médecine'),
    (re.compile(r'SPECIALT|SPECIALIT', re.I),           'Spécialité',                'TRAVAILLEUR', None),
    (re.compile(r'PROFESSIONNEL', re.I),                'Les Professionnels',        'TRAVAILLEUR', None),
    (re.compile(r'DOCTORAL|DOCTORAT', re.I),            'Cycle Doctoral',            'ETUDIANT', 'Doctorat'),
    (re.compile(r'\bMASTER\b', re.I),                   'Master',                    'ETUDIANT', 'Master'),
    (re.compile(r'INGENI', re.I),                       "Cycle d'Ingénierie",        'ETUDIANT', 'Cycle ingénieur'),
    (re.compile(r'\bLICENCE\b', re.I),                  'Licence',                   'ETUDIANT', 'Licence'),
]

_KNOWN_CAT_LABELS = frozenset(d[1] for d in _CAT_DEFS)


def _normalize_cat(raw):
    """
    Retourne (label_canonique, statut, cycle_nom) pour un en-tête de catégorie.
    Retourne (None, 'AUTRE', None) si non reconnu.
    """
    if not raw:
        return None, 'AUTRE', None
    plain = _plain(str(raw).strip())
    for pattern, label, statut, cycle in _CAT_DEFS:
        if pattern.search(plain):
            return label, statut, cycle
    return None, 'AUTRE', None


def _map_region(raw: str) -> str:
    """Mappe n'importe quel libellé de région au code CELLULE_VALID."""
    if not raw:
        return 'AUTRE'
    plain = _plain(re.sub(r'(?i)^\s*cellule\s*', '', str(raw)).strip())
    for key, val in CELLULE_MAP.items():
        if key in plain:
            return val
    return 'AUTRE'


def _clean_val(raw):
    """Normalise une valeur de cellule ; retourne None si vide."""
    if raw is None:
        return None
    s = str(raw).replace('\n', ' ').replace('\t', ' ').strip()
    s = re.sub(r'\s+', ' ', s)
    return s or None


_SKIP_EXACT = frozenset(_plain(w) for w in [
    'ABAGUMYABANGA', 'SECTION MAROC', 'SECTION', 'MAROC',
    'LES PROFESSIONNELS', 'PROFESSIONNELS', 'SPECIALTE', 'SPECIALITE',
    'SPECIALITE', 'SPECIALTÉ', 'N°', 'NOM', 'PRENOM', 'PRÉNOM',
    'LISTE', 'CYCLE DOCTORAL', 'DOCTORAL', 'DOCTORAT', 'MASTER',
    'INGENIEURIE', 'MEDECINE', 'LICENCE', 'CYCLE', 'CATEGORIE',
    'STATUT', 'CELLULE', 'NOM COMPLET', 'NOM ET PRENOM',
])


def _is_person_name(text: str) -> bool:
    """Heuristique : ce texte ressemble-t-il à un nom de personne ?"""
    if not text:
        return False
    text = text.strip()
    if len(text) < 3 or len(text) > 70:
        return False
    plain = _plain(text)
    if plain in _SKIP_EXACT:
        return False
    if plain.startswith('CELLULE'):
        return False
    # Si normalize_cat le reconnaît comme catégorie → pas un nom
    label, _, _ = _normalize_cat(text)
    if label:
        return False
    parts = text.split()
    if len(parts) < 2:
        return False
    alpha = sum(1 for c in text if c.isalpha() or c in " -'")
    if len(text) > 0 and alpha / len(text) < 0.70:
        return False
    if re.search(r'\d{3,}', text):
        return False
    return True


def _parse_nom_prenom(text: str):
    """
    Extrait (nom, prenom) depuis un texte combiné.
    Logique : premier mot tout-en-majuscules ≥ 2 lettres = nom de famille.
    """
    if not text:
        return None, None
    # Retirer titres courants
    t = re.sub(
        r'^(Dr\.?\s+|S\.E\.?\s+|M\.\s+|Mme\.?\s+|Prof\.?\s+)',
        '', text.strip(), flags=re.IGNORECASE,
    )
    t = re.sub(r'\s*\(.*?\)', '', t).strip()
    parts = t.split()
    if len(parts) < 2:
        return None, None
    # Le nom de famille est en MAJUSCULES
    if parts[0] == parts[0].upper() and len(parts[0]) > 1 and parts[0].isalpha():
        return parts[0], ' '.join(parts[1:])
    # Sinon cherche le premier mot tout-en-maj dans la liste
    for i, p in enumerate(parts):
        if p == p.upper() and len(p) > 1 and p.isalpha():
            nom = p
            prenom_parts = parts[:i] + parts[i+1:]
            return nom, ' '.join(prenom_parts)
    # Fallback : premier mot = nom
    return parts[0], ' '.join(parts[1:])


def _infer_sexe(prenom: str) -> str:
    if not prenom:
        return 'M'
    first = prenom.split()[0].lower().strip('.,')
    return 'F' if first in FEMALE_FIRST_NAMES else 'M'


def _vmap(ws) -> dict:
    """
    Construit {(row, col): valeur} en propageant les cellules fusionnées.
    Nécessite que le workbook soit ouvert sans read_only=True.
    """
    vm: dict = {}
    try:
        for merged in ws.merged_cells.ranges:
            top_val = ws.cell(merged.min_row, merged.min_col).value
            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    vm[(r, c)] = top_val
    except Exception:
        pass  # read_only workbooks n'ont pas merged_cells
    for row in ws.iter_rows():
        for cell in row:
            key = (cell.row, cell.column)
            if key not in vm:
                vm[key] = cell.value
    return vm


# ─────────────────────────────────────────────────────────────────────────────
# STRATÉGIES DE PARSING EXCEL
# ─────────────────────────────────────────────────────────────────────────────

_FLAT_KEYS = {
    'NOM', 'PRENOM', 'PRÉNOM', 'SEXE', 'CELLULE',
    'CATEGORIE', 'CATÉGORIE', 'STATUT', 'EMAIL',
    'TELEPHONE', 'TÉLÉPHONE', 'VILLE', 'OBSERVATION',
}


def _try_flat_table(ws, vm: dict, sheet_name: str):
    """
    Stratégie 1 — tableau plat classique.
    Cherche une ligne d'en-têtes (Nom, Prénom, Sexe…) dans les 10 premières lignes.
    Chaque ligne suivante = un membre.
    """
    header_row = None
    col_map: dict = {}

    for row in range(1, min(11, ws.max_row + 1)):
        row_upper = {}
        for col in range(1, min(25, ws.max_column + 1)):
            v = _clean_val(vm.get((row, col)))
            if v:
                row_upper[col] = _plain(v)

        hits = sum(1 for v in row_upper.values() if any(k in v for k in _FLAT_KEYS))
        if hits >= 2:
            header_row = row
            for col, v in row_upper.items():
                if 'NOM COMPLET' in v or 'NOM ET PRENOM' in v:
                    col_map.setdefault('nom_complet', col)
                elif 'NOM' in v and 'PRENOM' not in v:
                    col_map.setdefault('nom', col)
                elif 'PRENOM' in v or 'PRÉNOM' in v:
                    col_map.setdefault('prenom', col)
                elif v == 'SEXE':
                    col_map.setdefault('sexe', col)
                elif 'CELLULE' in v:
                    col_map.setdefault('cellule', col)
                elif 'CATÉG' in v or 'CATEG' in v:
                    col_map.setdefault('categorie', col)
                elif 'STATUT' in v:
                    col_map.setdefault('statut', col)
                elif 'VILLE' in v:
                    col_map.setdefault('ville', col)
                elif 'EMAIL' in v or 'MAIL' in v:
                    col_map.setdefault('email', col)
                elif 'TEL' in v:
                    col_map.setdefault('telephone', col)
                elif 'OBSERV' in v or 'NOTE' in v:
                    col_map.setdefault('observations', col)
            break

    if header_row is None:
        return None
    has_name = 'nom_complet' in col_map or ('nom' in col_map and 'prenom' in col_map)
    if not has_name:
        return None

    records = []
    for row in range(header_row + 1, ws.max_row + 1):
        first = _clean_val(vm.get((row, 1)))
        if not first or first.startswith('('):
            continue

        if 'nom_complet' in col_map:
            raw = _clean_val(vm.get((row, col_map['nom_complet'])))
            nom, prenom = _parse_nom_prenom(raw) if raw else (None, None)
        else:
            nom = _clean_val(vm.get((row, col_map['nom'])))
            prenom = _clean_val(vm.get((row, col_map['prenom'])))

        if not nom or not prenom or len(nom) < 2 or len(prenom) < 2:
            continue

        # Sexe
        raw_sexe = _clean_val(vm.get((row, col_map['sexe']))) if 'sexe' in col_map else None
        sexe = 'F' if raw_sexe and raw_sexe.upper().startswith('F') else _infer_sexe(prenom)

        # Cellule
        raw_cell = _clean_val(vm.get((row, col_map['cellule']))) if 'cellule' in col_map else None
        if raw_cell:
            cellule = raw_cell.upper() if raw_cell.upper() in CELLULE_VALID else _map_region(raw_cell)
        else:
            cellule = 'AUTRE'

        # Catégorie / statut
        raw_cat = _clean_val(vm.get((row, col_map['categorie']))) if 'categorie' in col_map else None
        cat_aff = 'DIASPORA'
        statut = 'AUTRE'
        cycle_nom = None
        if raw_cat:
            cu = raw_cat.upper()
            if cu in {'ABAGUMYABANGA', 'SYMPATHISANT', 'DIASPORA'}:
                cat_aff = cu
            else:
                label, statut, cycle_nom = _normalize_cat(raw_cat)
                cat_aff = 'ABAGUMYABANGA' if label else 'DIASPORA'

        # Statut colonne dédiée (override)
        raw_statut = _clean_val(vm.get((row, col_map['statut']))) if 'statut' in col_map else None
        if raw_statut:
            su = _plain(raw_statut)
            if 'ETUDIANT' in su:
                statut = 'ETUDIANT'
            elif 'TRAVAILLEUR' in su or 'PROFESSIONNEL' in su:
                statut = 'TRAVAILLEUR'
            elif 'SANS' in su:
                statut = 'SANS_ACTIVITE'
            elif su in {'ETUDIANT', 'TRAVAILLEUR', 'SANS_ACTIVITE', 'AUTRE'}:
                statut = su

        records.append({
            'nom': nom, 'prenom': prenom, 'sexe': sexe,
            'cellule': cellule,
            'categorie_affiliation': cat_aff,
            'statut_socio_pro': statut,
            'cycle_nom': cycle_nom,
            'ville_residence': _clean_val(vm.get((row, col_map['ville']))) or '' if 'ville' in col_map else '',
            'email': _clean_val(vm.get((row, col_map['email']))) or '' if 'email' in col_map else '',
            'telephone': _clean_val(vm.get((row, col_map['telephone']))) or '' if 'telephone' in col_map else '',
            'observations': _clean_val(vm.get((row, col_map['observations']))) or '' if 'observations' in col_map else '',
            'source': f'{sheet_name}:flat_table',
        })

    return records or None


def _try_cross_table(ws, vm: dict, sheet_name: str):
    """
    Stratégie 2 — tableau croisé style ABAGUMYABANGA.
    Détecte des lignes où col A contient "CELLULE <région>".
    La première telle ligne avec des libellés de catégories en cols B-I établit
    la mise en page des colonnes ; les blocs suivants utilisent la même mise en page
    mais leurs lignes CELLULE peuvent aussi contenir des noms directement.
    """
    # Vérifier qu'au moins une ligne de col A contient "CELLULE"
    has_cellule_col_a = any(
        ws.cell(r, 1).value and 'CELLULE' in str(ws.cell(r, 1).value).upper()
        for r in range(1, ws.max_row + 1)
    )
    if not has_cellule_col_a:
        return None

    def _row_has_cat_labels(row_num):
        for col in range(2, 12):
            v = vm.get((row_num, col))
            if v:
                lbl, _, _ = _normalize_cat(v)
                if lbl in _KNOWN_CAT_LABELS:
                    return True
        return False

    records = []
    col_cats: dict = {}
    current_cellule = 'AUTRE'

    for row in range(1, ws.max_row + 1):
        direct_a = ws.cell(row, 1).value
        is_cellule_row = direct_a and 'CELLULE' in str(direct_a).upper()

        if is_cellule_row:
            current_cellule = _map_region(str(direct_a))
            if _row_has_cat_labels(row):
                # Ligne d'en-têtes pure : lire les catégories, ne pas extraire de noms
                col_cats = {}
                for col in range(2, 12):
                    v = vm.get((row, col))
                    if v:
                        lbl, stat, cyc = _normalize_cat(v)
                        if lbl:
                            col_cats[col] = (lbl, stat, cyc)
                continue
            # Sinon : la ligne CELLULE contient aussi des noms → tomber dans data

        if col_cats:
            for col, (lbl, stat, cyc) in col_cats.items():
                raw = _clean_val(vm.get((row, col)))
                if raw and _is_person_name(raw):
                    nom, prenom = _parse_nom_prenom(raw)
                    if nom and prenom and len(nom) >= 2 and len(prenom) >= 2:
                        records.append({
                            'nom': nom, 'prenom': prenom,
                            'sexe': _infer_sexe(prenom),
                            'cellule': current_cellule,
                            'categorie_affiliation': 'ABAGUMYABANGA',
                            'statut_socio_pro': stat,
                            'cycle_nom': cyc,
                            'ville_residence': '', 'email': '',
                            'telephone': '',
                            'observations': f'Catégorie : {lbl}',
                            'source': f'{sheet_name}:cross_table',
                        })

    return records or None


def _try_block_table(ws, vm: dict, sheet_name: str):
    """
    Stratégie 3 — blocs dispersés (style Feuil2 ABAGUMYABANGA).
    Détecte automatiquement les lignes qui contiennent ≥ 2 libellés de catégories
    → ce sont les lignes d'en-têtes de blocs.
    Remonte jusqu'à 10 lignes pour trouver le libellé de région.
    Extrait les noms des lignes de données suivantes.
    """
    # Identifier les lignes d'en-têtes de blocs (≥ 2 catégories dans la même ligne)
    header_rows = []
    for row in range(1, ws.max_row + 1):
        cat_count = 0
        for col in range(1, min(ws.max_column + 1, 25)):
            v = vm.get((row, col))
            if v:
                lbl, _, _ = _normalize_cat(v)
                if lbl:
                    cat_count += 1
        if cat_count >= 2:
            header_rows.append(row)

    if not header_rows:
        return None

    records = []

    for hi, hrow in enumerate(header_rows):
        # Lire les catégories de cette ligne d'en-têtes
        col_cats: dict = {}
        for col in range(1, min(ws.max_column + 1, 25)):
            v = vm.get((hrow, col))
            if v:
                lbl, stat, cyc = _normalize_cat(v)
                if lbl:
                    col_cats[col] = (lbl, stat, cyc)

        if not col_cats:
            continue

        # Rechercher le libellé de région en remontant (dans la plage de colonnes du bloc)
        cat_cols = set(col_cats.keys())
        min_col, max_col = min(cat_cols), max(cat_cols)
        region_label = None
        cellule_code = 'AUTRE'

        for lookback in range(1, 11):
            r = hrow - lookback
            if r < 1:
                break
            for col in range(max(1, min_col - 2), max_col + 3):
                v = _clean_val(vm.get((r, col)))
                if v and 2 < len(v) <= 60:
                    # Ne doit pas être lui-même un label de catégorie
                    lbl, _, _ = _normalize_cat(v)
                    if lbl:
                        continue
                    # Ne doit pas être un nom de personne
                    if _is_person_name(v):
                        continue
                    region_label = re.sub(r'(?i)^\s*cellule\s*', '', v).strip()
                    cellule_code = _map_region(v)
                    break
            if region_label:
                break

        # Plage de données : de hrow+1 jusqu'à la prochaine ligne d'en-têtes
        next_hrow = header_rows[hi + 1] if hi + 1 < len(header_rows) else ws.max_row + 1

        for row in range(hrow + 1, next_hrow):
            for col, (lbl, stat, cyc) in col_cats.items():
                raw = _clean_val(vm.get((row, col)))
                if raw and _is_person_name(raw):
                    nom, prenom = _parse_nom_prenom(raw)
                    if nom and prenom and len(nom) >= 2 and len(prenom) >= 2:
                        records.append({
                            'nom': nom, 'prenom': prenom,
                            'sexe': _infer_sexe(prenom),
                            'cellule': cellule_code,
                            'categorie_affiliation': 'ABAGUMYABANGA',
                            'statut_socio_pro': stat,
                            'cycle_nom': cyc,
                            'ville_residence': '', 'email': '',
                            'telephone': '',
                            'observations': f'Région : {region_label or "?"} | Catégorie : {lbl}',
                            'source': f'{sheet_name}:block_table',
                        })

    return records or None


def _parse_excel_workbook(wb) -> tuple:
    """
    Point d'entrée Excel : essaie les 3 stratégies sur chaque feuille.
    Retourne (records, strategies_used).
    """
    _SKIP_SHEETS = {'données', 'donnees', 'output', 'résultats', 'resultats', 'export'}
    records = []
    strategies = []

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in _SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row < 2:
            continue
        vm = _vmap(ws)

        for fn, name in [
            (_try_flat_table,  'flat_table'),
            (_try_cross_table, 'cross_table'),
            (_try_block_table, 'block_table'),
        ]:
            result = fn(ws, vm, sheet_name)
            if result:
                records.extend(result)
                strategies.append(f'{sheet_name}:{name}')
                break  # une seule stratégie par feuille

    return records, strategies


# ─────────────────────────────────────────────────────────────────────────────
# CRÉATION DE MEMBRES (partagée PDF + Excel)
# ─────────────────────────────────────────────────────────────────────────────

@transaction.atomic
def _create_membres(records: list) -> dict:
    """Insère les membres et profils étudiants à partir d'une liste de dicts."""
    created = skipped = 0
    errors = []
    _cache: dict = {}

    def _cycle(nom):
        if nom not in _cache.setdefault('cyc', {}):
            obj, _ = Cycle.objects.get_or_create(nom=nom)
            _cache['cyc'][nom] = obj
        return _cache['cyc'][nom]

    def _domaine():
        key = 'Non spécifié'
        if key not in _cache.setdefault('dom', {}):
            obj, _ = Domaine.objects.get_or_create(nom=key)
            _cache['dom'][key] = obj
        return _cache['dom'][key]

    def _filiere(dom):
        if dom.id not in _cache.setdefault('fil', {}):
            obj, _ = Filiere.objects.get_or_create(nom='Non spécifiée', domaine=dom)
            _cache['fil'][dom.id] = obj
        return _cache['fil'][dom.id]

    def _niveau():
        key = 'Non spécifié'
        if key not in _cache.setdefault('niv', {}):
            obj, _ = Niveau.objects.get_or_create(nom=key)
            _cache['niv'][key] = obj
        return _cache['niv'][key]

    for i, r in enumerate(records):
        try:
            membre, was_created = Membre.objects.get_or_create(
                nom__iexact=r['nom'],
                prenom__iexact=r['prenom'],
                defaults={
                    'nom': r['nom'],
                    'prenom': r['prenom'],
                    'sexe': r.get('sexe', 'M'),
                    'cellule': r.get('cellule', ''),
                    'statut_socio_pro': r.get('statut_socio_pro', 'AUTRE'),
                    'categorie_affiliation': r.get('categorie_affiliation', 'DIASPORA'),
                    'statut_compte': 'ACTIF',
                    'ville_residence': r.get('ville_residence', ''),
                    'email': r.get('email', ''),
                    'telephone': r.get('telephone', ''),
                    'observations': r.get('observations', 'Importé automatiquement'),
                },
            )
            if was_created:
                created += 1
                cycle_nom = r.get('cycle_nom')
                if r.get('statut_socio_pro') == 'ETUDIANT' and cycle_nom:
                    dom = _domaine()
                    ProfilEtudiant.objects.get_or_create(
                        membre=membre,
                        defaults={
                            'cycle': _cycle(cycle_nom),
                            'domaine': dom,
                            'filiere': _filiere(dom),
                            'niveau': _niveau(),
                            'etablissement': 'Non spécifié',
                            'ville_etudes': '',
                            'annee_academique': '2025-2026',
                        },
                    )
            else:
                skipped += 1
        except Exception as e:
            errors.append({'index': i + 1, 'nom': r.get('nom', '?'), 'erreur': str(e)})

    return {'crees': created, 'ignores': skipped, 'erreurs': errors}


# ─────────────────────────────────────────────────────────────────────────────
# VIEWSET
# ─────────────────────────────────────────────────────────────────────────────

class MembreViewSet(viewsets.ModelViewSet):
    queryset = Membre.objects.select_related("profil_etudiant").prefetch_related("cotisations")
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = MembreFilter
    search_fields = ["nom", "prenom", "email", "telephone", "ville_residence"]
    ordering_fields = ["nom", "prenom", "date_adhesion", "categorie_affiliation", "cellule"]
    ordering = ["nom"]

    def get_serializer_class(self):
        if self.action == "list":
            return MembreListSerializer
        if self.action in ["create", "update", "partial_update"]:
            return MembreCreateUpdateSerializer
        return MembreDetailSerializer

    @action(detail=False, methods=["get"], url_path="villes")
    def villes(self, request):
        return Response(VILLES_MAROC)

    @action(detail=False, methods=["get"], url_path="cellules")
    def cellules(self, request):
        return Response([{"value": k, "label": v} for k, v in Membre.CELLULE_CHOICES])

    # ── Export Excel ──────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Membres"

        hf = Font(bold=True, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="CE1126")
        headers = [
            "ID", "Nom", "Prénom", "Sexe", "Cellule", "Catégorie", "Statut socio-pro",
            "Statut compte", "Ville", "Email", "Téléphone", "Date adhésion",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")

        for row, m in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=m.id)
            ws.cell(row=row, column=2, value=m.nom)
            ws.cell(row=row, column=3, value=m.prenom)
            ws.cell(row=row, column=4, value=m.get_sexe_display())
            ws.cell(row=row, column=5, value=m.get_cellule_display() if m.cellule else "")
            ws.cell(row=row, column=6, value=m.get_categorie_affiliation_display())
            ws.cell(row=row, column=7, value=m.get_statut_socio_pro_display())
            ws.cell(row=row, column=8, value=m.get_statut_compte_display())
            ws.cell(row=row, column=9, value=m.ville_residence)
            ws.cell(row=row, column=10, value=m.email)
            ws.cell(row=row, column=11, value=m.telephone)
            ws.cell(row=row, column=12, value=str(m.date_adhesion))

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="membres_cndd_fdd.xlsx"'
        wb.save(response)
        return response

    # ── Modèle d'import ───────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="template-import")
    def template_import(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Membres"

        hf = Font(bold=True, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="CE1126")
        headers = [
            "Nom *", "Prénom *", "Sexe *", "Catégorie *", "Cellule",
            "Statut socio-pro", "Statut compte", "Ville résidence",
            "Téléphone", "Email", "Date naissance", "Observations",
        ]
        notes = [
            "", "", "M ou F",
            "ABAGUMYABANGA / SYMPATHISANT / DIASPORA",
            "TANGER_TETOUAN_OUJDA / KENITRA / AGADIR / RABAT_SALE / LAAYOUNE_DAKHLA / FEZ_MEKNES / CASABLANCA",
            "ETUDIANT / TRAVAILLEUR / SANS_ACTIVITE / AUTRE",
            "ACTIF / INACTIF / SUSPENDU",
            "", "", "", "Format: AAAA-MM-JJ", "",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")

        ws.append([
            "NIYONKURU", "Jean", "M", "ABAGUMYABANGA", "RABAT_SALE", "ETUDIANT", "ACTIF",
            "Rabat", "+212612345678", "jean@email.com", "2000-01-15", "Exemple",
        ])
        ws.append(["" if not n else f"({n})" for n in notes])
        for col, note in enumerate(notes, 1):
            if note:
                ws.cell(row=3, column=col).font = Font(italic=True, color="999999", size=8)

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="modele_import_membres.xlsx"'
        wb.save(response)
        return response

    # ── Import Excel — auto-détection de structure ────────────────────────────

    @action(detail=False, methods=["post"], url_path="import-excel",
            parser_classes=[MultiPartParser])
    def import_excel(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Aucun fichier fourni"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # data_only=True pour résoudre les formules ; PAS read_only pour merged_cells
            wb = openpyxl.load_workbook(file, data_only=True)
        except Exception:
            return Response({"error": "Fichier Excel invalide (.xlsx requis)"},
                            status=status.HTTP_400_BAD_REQUEST)

        records, strategies = _parse_excel_workbook(wb)

        if not records:
            return Response({
                "error": (
                    "Aucune donnée exploitable trouvée. "
                    "Formats supportés : "
                    "(1) tableau plat avec colonnes Nom/Prénom, "
                    "(2) tableau croisé CELLULE × catégories (col A = CELLULE), "
                    "(3) blocs région + en-têtes de catégories."
                )
            }, status=status.HTTP_400_BAD_REQUEST)

        result = _create_membres(records)
        result['extraits'] = len(records)
        result['strategies'] = strategies
        return Response(result)

    # ── Import PDF ────────────────────────────────────────────────────────────

    @action(detail=False, methods=["post"], url_path="import-pdf",
            parser_classes=[MultiPartParser])
    def import_pdf(self, request):
        try:
            import pdfplumber
        except ImportError:
            return Response(
                {"error": "Le module pdfplumber n'est pas installé sur le serveur."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Aucun fichier fourni"}, status=status.HTTP_400_BAD_REQUEST)

        def _classify_pdf_header(text_upper):
            """Retourne (statut, cycle) ou (None, None) si ce n'est pas un en-tête."""
            lbl, stat, cyc = _normalize_cat(text_upper)
            if lbl:
                return stat, cyc
            return None, None

        extracted = []
        current_cellule = ''
        current_statut = 'AUTRE'
        current_cycle = None
        col_info: dict = {}  # persiste d'une page à l'autre (tableaux qui se coupent)

        try:
            with pdfplumber.open(file) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    if not tables:
                        continue

                    for table in tables:
                        for row in table:
                            if not row:
                                continue

                            # Passe 0 : détecter les CELLULE → mettre à jour current_cellule
                            # (séparé de la détection d'en-têtes car une ligne peut avoir
                            #  CELLULE + des noms dans d'autres colonnes)
                            for ci, cell in enumerate(row):
                                if not cell:
                                    continue
                                upper = str(cell).strip().upper()
                                if 'CELLULE' in upper:
                                    detected = _map_region(upper)
                                    if detected != 'AUTRE':
                                        current_cellule = detected

                            # Passe 1 : détecter les lignes d'en-têtes de catégories
                            # Une ligne est un en-tête seulement si elle contient de vraies
                            # catégories (pas juste CELLULE).
                            new_col_info = {}
                            header_found = False
                            for ci, cell in enumerate(row):
                                if not cell:
                                    continue
                                upper = str(cell).strip().upper()
                                if 'CELLULE' in upper:
                                    new_col_info[ci] = ('SKIP', None)
                                    # Ne PAS déclencher header_found pour CELLULE
                                else:
                                    stat, cyc = _classify_pdf_header(upper)
                                    if stat is not None:
                                        new_col_info[ci] = (stat, cyc)
                                        header_found = True

                            if header_found:
                                col_info = new_col_info
                                # Si ci=0 porte une catégorie académique (pas CELLULE) →
                                # section étudiants listés sans info cellule dans ce PDF
                                c0 = new_col_info.get(0)
                                if c0 and c0[0] not in ('SKIP', None):
                                    current_cellule = ''
                                for v in col_info.values():
                                    if v[0] not in ('SKIP', None):
                                        current_statut = v[0]
                                        current_cycle = v[1]
                                        break
                                continue

                            # Passe 2 : extraire les noms
                            for ci, cell in enumerate(row):
                                if not cell:
                                    continue
                                cs = str(cell).strip()
                                if not cs:
                                    continue

                                if 'CELLULE' in cs.upper():
                                    continue  # déjà traité en passe 0

                                col_meta = col_info.get(ci)
                                if col_meta and col_meta[0] == 'SKIP':
                                    continue

                                cell_statut = (col_meta[0] if col_meta else None) or current_statut
                                cell_cycle = (col_meta[1] if col_meta else None) or current_cycle

                                if _is_person_name(cs):
                                    nom, prenom = _parse_nom_prenom(cs)
                                    if nom and prenom and len(nom) >= 2 and len(prenom) >= 2:
                                        extracted.append({
                                            'nom': nom, 'prenom': prenom,
                                            'sexe': _infer_sexe(prenom),
                                            'cellule': current_cellule,
                                            'statut_socio_pro': cell_statut,
                                            'cycle_nom': cell_cycle,
                                            'categorie_affiliation': 'ABAGUMYABANGA',
                                            'observations': 'Importé depuis PDF ABAGUMYABANGA',
                                        })

        except Exception as e:
            return Response(
                {"error": f"Erreur lecture PDF : {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        result = _create_membres(extracted)
        result['extraits'] = len(extracted)
        return Response(result)

    # ── Étudiants ─────────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="etudiants")
    def etudiants(self, request):
        queryset = self.filter_queryset(
            self.get_queryset().filter(statut_socio_pro="ETUDIANT")
        )
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = MembreDetailSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        return Response(MembreDetailSerializer(queryset, many=True).data)


class ProfilEtudiantViewSet(viewsets.ModelViewSet):
    queryset = ProfilEtudiant.objects.select_related(
        "membre", "cycle", "domaine", "filiere", "niveau"
    ).all()
    serializer_class = ProfilEtudiantSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ["cycle", "domaine", "filiere", "niveau", "annee_academique", "statut_parcours"]
    search_fields = ["membre__nom", "membre__prenom", "etablissement", "ville_etudes"]
